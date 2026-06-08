"""Reports — PDF (Russian) and XLSX export."""
import io, os, logging
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from app.database import get_db
from app.models import User, UserRole, Interview, Task, Submission, AnticheatEvent
from app.auth.dependencies import require_role
logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/hr/report", tags=["Reports"])

def _font():
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    for r, b in [("C:/Windows/Fonts/arial.ttf","C:/Windows/Fonts/arialbd.ttf"),("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf","/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf")]:
        if os.path.exists(r):
            try:
                pdfmetrics.registerFont(TTFont('F', r))
                pdfmetrics.registerFont(TTFont('FB', b if os.path.exists(b) else r))
                return 'F', 'FB'
            except: continue
    return 'Helvetica', 'Helvetica-Bold'

@router.get("/session/{sid}/pdf")
async def export_pdf(sid: int, current_user: User = Depends(require_role([UserRole.HR])), db: AsyncSession = Depends(get_db)):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    f, fb = _font()
    iv = (await db.execute(select(Interview).where(Interview.id == sid))).scalar_one_or_none()
    if not iv: raise HTTPException(404)
    u = (await db.execute(select(User).where(User.id == iv.user_id))).scalar_one_or_none()
    tasks = (await db.execute(select(Task).where(Task.interview_id == sid).order_by(Task.order_number))).scalars().all()
    ac_r = await db.execute(select(func.sum(AnticheatEvent.severity)).where(AnticheatEvent.interview_id == sid))
    susp = min((ac_r.scalar() or 0) / 5.0, 1.0)
    acs = (await db.execute(select(AnticheatEvent).where(AnticheatEvent.interview_id == sid))).scalars().all()

    ts = ParagraphStyle('t', fontName=fb, fontSize=18, spaceAfter=20, alignment=1)
    hs = ParagraphStyle('h', fontName=fb, fontSize=13, spaceAfter=8, spaceBefore=15, textColor=colors.HexColor('#1e40af'))
    ns = ParagraphStyle('n', fontName=f, fontSize=10, spaceAfter=5)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
    el = []
    el.append(Paragraph("Отчёт о техническом собеседовании", ts))
    rec_map = {'hire':'Нанять','maybe':'Возможно','reject':'Отказ'}
    info = [["ФИО:", u.full_name if u else "—"],["Email:", u.email if u else "—"],["Уровень:", iv.level.value.upper()],
            ["Дата:", iv.started_at.strftime("%d.%m.%Y %H:%M") if iv.started_at else "—"],["Статус:", "Завершено" if iv.status.value=="completed" else iv.status.value],
            ["Общий балл:", f"{iv.total_score:.1f}%"],["Рекомендация:", rec_map.get(iv.ai_recommendation,"Нет данных")],["Подозрительность:", f"{susp:.0%}"]]
    t = Table(info, colWidths=[4.5*cm, 11*cm])
    t.setStyle(TableStyle([('FONTNAME',(0,0),(0,-1),fb),('FONTNAME',(1,0),(1,-1),f),('FONTSIZE',(0,0),(-1,-1),10),('BOTTOMPADDING',(0,0),(-1,-1),5)]))
    el.append(t); el.append(Spacer(1,15))
    el.append(Paragraph("Результаты по задачам", hs))
    for task in tasks:
        el.append(Paragraph(f"Задача {task.order_number}: {task.title}", ParagraphStyle('tt',fontName=fb,fontSize=11,spaceAfter=4,spaceBefore=10)))
        sub = (await db.execute(select(Submission).where(Submission.task_id==task.id).order_by(Submission.score.desc()).limit(1))).scalar_one_or_none()
        if sub:
            td = [["Балл:", f"{sub.score:.1f}%"],["Видимые тесты:", f"{sub.passed_visible}/{sub.total_visible}"],["Скрытые тесты:", f"{sub.passed_hidden}/{sub.total_hidden}"],["Язык:", sub.language]]
            tt = Table(td, colWidths=[4.5*cm,11*cm]); tt.setStyle(TableStyle([('FONTNAME',(0,0),(-1,-1),f),('FONTSIZE',(0,0),(-1,-1),9)])); el.append(tt)
        else: el.append(Paragraph("Решение не отправлено", ns))
    if acs:
        el.append(Spacer(1,12)); el.append(Paragraph("Античит-события", hs))
        for ev in acs:
            labels={'paste':'Вставка кода','tab_switch':'Переключение вкладки','devtools':'DevTools','copy':'Копирование','rapid_code':'Быстрая вставка'}
            el.append(Paragraph(f"• {labels.get(ev.event_type,ev.event_type)}", ns))
    if iv.ai_summary: el.append(Spacer(1,12)); el.append(Paragraph("Итоговая оценка", hs)); el.append(Paragraph(iv.ai_summary, ns))
    doc.build(el); buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="report_{sid}.pdf"'})

@router.get("/export/xlsx")
async def export_xlsx(current_user: User = Depends(require_role([UserRole.HR])), db: AsyncSession = Depends(get_db)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook(); ws = wb.active; ws.title = "Кандидаты"
    headers = ["ID","ФИО","Email","Телефон","Сессий","Ср. балл","Уровень","Подозрительность","Решение","Дата"]
    hf = PatternFill(start_color="2E75B6",end_color="2E75B6",fill_type="solid"); hfont = Font(color="FFFFFF",bold=True)
    bd = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    for c,h in enumerate(headers,1): cell=ws.cell(row=1,column=c,value=h); cell.fill=hf; cell.font=hfont; cell.border=bd
    users = (await db.execute(select(User).where(User.role==UserRole.CANDIDATE))).scalars().all()
    for ri,u in enumerate(users,2):
        sr = await db.execute(select(func.count(Interview.id),func.avg(Interview.total_score),func.max(Interview.started_at)).where(Interview.user_id==u.id))
        s=sr.one(); li=(await db.execute(select(Interview).where(Interview.user_id==u.id).order_by(Interview.started_at.desc()).limit(1))).scalar_one_or_none()
        susp=0.0
        if li:
            sr2=await db.execute(select(func.sum(AnticheatEvent.severity)).where(AnticheatEvent.interview_id==li.id)); susp=min((sr2.scalar() or 0)/5.0,1.0)
        row=[u.id,u.full_name,u.email,u.phone or"",s[0]or 0,round(float(s[1]or 0),2),li.level.value.upper()if li else"",f"{susp:.0%}",
             {'hire':'Нанять','maybe':'Возможно','reject':'Отказ'}.get(li.ai_recommendation,"")if li else"",s[2].strftime("%d.%m.%Y")if s[2]else""]
        for c,v in enumerate(row,1): cell=ws.cell(row=ri,column=c,value=v); cell.border=bd
    for col in ws.columns: ws.column_dimensions[col[0].column_letter].width=min(max(len(str(c.value or""))for c in col)+2,30)
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="candidates.xlsx"'})
